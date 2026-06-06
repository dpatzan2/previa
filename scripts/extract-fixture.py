from __future__ import annotations

import json
from datetime import datetime, time
from pathlib import Path
from unicodedata import normalize

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path.home() / "Downloads" / "Fixture-Copa-Mundial-FIFA-2026_ClasesExcel.xlsx"
OUT = ROOT / "prisma" / "fixture-data.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return " ".join(value.strip().split())
    return value


def normalize_team(name: str) -> str:
    text = normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.upper().replace(".", "").split())


def stage_for(match_number: int) -> str:
    if match_number <= 72:
        return "GROUP"
    if match_number <= 88:
        return "ROUND_OF_32"
    if match_number <= 96:
        return "ROUND_OF_16"
    if match_number <= 100:
        return "QUARTER_FINAL"
    if match_number <= 102:
        return "SEMIFINAL"
    if match_number == 103:
        return "THIRD_PLACE"
    return "FINAL"


def as_iso(date_value, time_value):
    if not isinstance(date_value, datetime):
        return None
    hour = minute = 0
    if isinstance(time_value, time):
        hour = time_value.hour
        minute = time_value.minute
    return datetime(date_value.year, date_value.month, date_value.day, hour, minute).strftime(
        "%Y-%m-%dT%H:%M:%S-06:00"
    )


def main():
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb["Recursos"]
    table = ws.tables["tDatos"]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [ws.cell(min_row, col).value for col in range(min_col, max_col + 1)]

    matches = []
    teams = {}
    for row in range(min_row + 1, max_row + 1):
        record = {
            headers[index]: clean(ws.cell(row, min_col + index).value)
            for index in range(len(headers))
        }
        number = record["#"]
        if not isinstance(number, int):
            continue
        raw_group = record.get("GRUPO")
        group_code = None
        if isinstance(raw_group, str) and raw_group.startswith("Grupo "):
            group_code = raw_group.replace("Grupo ", "").strip()

        home = record.get("EQUIPO 1")
        away = record.get("EQUIPO 2")
        for team in (home, away):
            if isinstance(team, str) and team:
                teams[normalize_team(team)] = {"name": team, "groupCode": group_code}

        matches.append(
            {
                "matchNumber": number,
                "stage": stage_for(number),
                "groupCode": group_code,
                "dateLabel": record.get("FECHA LARGA"),
                "timeLabel": record.get("HORA").strftime("%H:%M")
                if isinstance(record.get("HORA"), time)
                else None,
                "kickoffAt": as_iso(record.get("FECHA"), record.get("HORA")),
                "venue": record.get("ESTADIO"),
                "venueShort": record.get("ESTADIO ABREV."),
                "homeTeam": home if isinstance(home, str) and home else None,
                "awayTeam": away if isinstance(away, str) and away else None,
                "homePlaceholder": home if isinstance(home, str) and home else f"Local P{number}",
                "awayPlaceholder": away if isinstance(away, str) and away else f"Visitante P{number}",
            }
        )

    OUT.write_text(
        json.dumps(
            {"teams": list(teams.values()), "matches": matches},
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(matches)} matches and {len(teams)} teams to {OUT}")


if __name__ == "__main__":
    main()
